"""Excel import utilities for customer bulk import."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.customers.models import Customer
from apps.customers.serializers import CustomerImportSerializer
from apps.reps.models import Rep


# Template column headers (order matters)
TEMPLATE_HEADERS = [
    "name",
    "phone",
    "email",
    "assigned_rep_codes",
]

# Arabic labels for template
TEMPLATE_LABELS = {
    "name": "الاسم *",
    "phone": "رقم الهاتف *",
    "email": "البريد الإلكتروني",
    "assigned_rep_codes": "أكواد المندوبين (مفصولة بفاصلة)",
}

# Sample data for template
SAMPLE_DATA = [
    {
        "name": "أحمد محمود",
        "phone": "+963991234567",
        "email": "ahmad@example.com",
        "assigned_rep_codes": "REP001,REP002",
    },
    {
        "name": "فاطمة علي",
        "phone": "+963992345678",
        "email": "",
        "assigned_rep_codes": "REP001",
    },
]


def generate_template() -> BytesIO:
    """
    Generate an Excel template file with headers and sample data.
    
    Returns:
        BytesIO: Excel file in memory
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Customers Template"
    
    # Style for headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Write headers
    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = TEMPLATE_LABELS.get(header, header)
        cell.fill = header_fill
        cell.font = header_font
    
    # Write sample data
    for row_idx, sample in enumerate(SAMPLE_DATA, start=2):
        for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = sample.get(header, "")
    
    # Auto-adjust column widths
    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 20
    
    # Add notes sheet
    notes_ws = wb.create_sheet("ملاحظات")
    notes_ws.append(["ملاحظات مهمة حول تنسيق الملف:"])
    notes_ws.append([])
    notes_ws.append(["1. الحقول المطلوبة (يجب ملؤها):"])
    notes_ws.append(["   - الاسم"])
    notes_ws.append(["   - رقم الهاتف (بصيغة: +963XXXXXXXXX)"])
    notes_ws.append([])
    notes_ws.append(["2. الحقول الاختيارية:"])
    notes_ws.append(["   - البريد الإلكتروني"])
    notes_ws.append(["   - أكواد المندوبين (مفصولة بفاصلة مثل: REP001,REP002)"])
    notes_ws.append([])
    notes_ws.append(["3. ملاحظات:"])
    notes_ws.append(["   - التصنيف والموقع سيتم تحديدهما يدوياً لاحقاً"])
    notes_ws.append(["   - قم بحذف الصفوف النموذجية وأضف بياناتك الخاصة"])
    notes_ws.append(["   - لا تقم بتغيير عناوين الأعمدة في السطر الأول"])
    notes_ws.append(["   - يمكنك إضافة عدد غير محدود من الصفوف"])
    
    notes_ws.column_dimensions['A'].width = 80
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def parse_excel_file(file, company_id: int) -> dict[str, Any]:
    """
    Parse uploaded Excel file and import customers.
    
    Args:
        file: Uploaded file object from request.FILES
        company_id: ID of the company importing customers
    
    Returns:
        Dictionary with import results:
        {
            "total_rows": int,
            "successful": int,
            "failed": int,
            "created_customers": [...],
            "errors": [
                {"row": int, "data": {...}, "errors": {...}},
                ...
            ]
        }
    """
    try:
        # Load workbook
        wb = load_workbook(filename=file, read_only=True, data_only=True)
        ws = wb.active
        
        # Get all rows (skip header)
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        if not rows:
            return {
                "total_rows": 0,
                "successful": 0,
                "failed": 0,
                "created_customers": [],
                "errors": [{"row": 0, "error": "الملف فارغ أو لا يحتوي على بيانات"}],
            }
        
        # Get company's active reps for validation
        company_reps = Rep.objects.filter(company_id=company_id, is_active=True)
        rep_code_to_id = {rep.referral_code: rep.id for rep in company_reps}
        
        # Get existing customer phones to check for duplicates
        existing_phones = set(
            Customer.objects.values_list("phone", flat=True)
        )
        
        results = {
            "total_rows": len(rows),
            "successful": 0,
            "failed": 0,
            "created_customers": [],
            "errors": [],
        }
        
        # Process each row
        for row_idx, row in enumerate(rows, start=2):  # Start at 2 (after header)
            try:
                # Map row to dictionary
                row_data = {}
                for col_idx, header in enumerate(TEMPLATE_HEADERS):
                    value = row[col_idx] if col_idx < len(row) else None
                    # Convert to string and strip, handle None
                    if value is not None:
                        row_data[header] = str(value).strip() if str(value).strip() else None
                    else:
                        row_data[header] = None
                
                # Skip completely empty rows
                if all(v is None or v == "" for v in row_data.values()):
                    continue
                
                # Validate row data
                serializer = CustomerImportSerializer(data=row_data)
                
                if not serializer.is_valid():
                    results["failed"] += 1
                    results["errors"].append({
                        "row": row_idx,
                        "data": row_data,
                        "errors": serializer.errors,
                    })
                    continue
                
                validated_data = serializer.validated_data
                
                # Check if phone already exists
                if validated_data["phone"] in existing_phones:
                    results["failed"] += 1
                    results["errors"].append({
                        "row": row_idx,
                        "data": row_data,
                        "errors": {"phone": ["رقم الهاتف مستخدم من قبل"]},
                    })
                    continue
                
                # Process assigned_rep_codes
                rep_ids = []
                if validated_data.get("assigned_rep_codes"):
                    codes = [
                        code.strip()
                        for code in validated_data["assigned_rep_codes"].split(",")
                        if code.strip()
                    ]
                    
                    invalid_codes = []
                    for code in codes:
                        if code in rep_code_to_id:
                            rep_ids.append(rep_code_to_id[code])
                        else:
                            invalid_codes.append(code)
                    
                    if invalid_codes:
                        results["failed"] += 1
                        results["errors"].append({
                            "row": row_idx,
                            "data": row_data,
                            "errors": {
                                "assigned_rep_codes": [
                                    f"الأكواد التالية غير صحيحة أو غير نشطة: {', '.join(invalid_codes)}"
                                ]
                            },
                        })
                        continue
                
                # Create customer
                customer_data = {
                    "name": validated_data["name"],
                    "phone": validated_data["phone"],
                    "email": validated_data.get("email") or None,
                    "category": validated_data.get("category") or None,
                    "latitude": validated_data.get("latitude"),
                    "longitude": validated_data.get("longitude"),
                    "is_active": True,
                }
                
                customer = Customer.objects.create(**customer_data)
                
                # Assign reps if any
                if rep_ids:
                    customer.assigned_reps.set(rep_ids)
                
                # Add to existing phones to prevent duplicates within same import
                existing_phones.add(validated_data["phone"])
                
                results["successful"] += 1
                results["created_customers"].append({
                    "id": customer.id,
                    "name": customer.name,
                    "phone": customer.phone,
                    "row": row_idx,
                })
                
            except Exception as e:
                results["failed"] += 1
                results["errors"].append({
                    "row": row_idx,
                    "data": row_data if 'row_data' in locals() else {},
                    "errors": {"general": [f"خطأ غير متوقع: {str(e)}"]},
                })
        
        return results
        
    except Exception as e:
        return {
            "total_rows": 0,
            "successful": 0,
            "failed": 0,
            "created_customers": [],
            "errors": [{"row": 0, "error": f"فشل في معالجة الملف: {str(e)}"}],
        }
